#!/usr/bin/env python3
"""
BIZRA Lexicon Ledger Builder (v0.2)

Inputs (expected in the same folder unless overridden via args):
- bizra_term_frequencies_per_model.csv
- combined_concept_graph.json
- combined_concept_graph_edges.csv
- monthly_theme_trends_all_models.csv
- conversation_metrics_all_models.csv

Output:
- BIZRA_Lexicon_Ledger_v0.2.xlsx

This script is designed to be deterministic and safe:
- No network access required
- No destructive writes (writes a new .xlsx)
"""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def entropy(vec: np.ndarray) -> float:
    s = float(vec.sum())
    if s <= 0:
        return 0.0
    p = vec / s
    p = p[p > 0]
    return float(-(p * np.log(p)).sum())

def drift_action(total: float, presence: int, cv: float, max_share: float) -> str:
    cv = 0.0 if (cv is None or np.isnan(cv)) else float(cv)
    max_share = 0.0 if (max_share is None or np.isnan(max_share)) else float(max_share)

    if presence == 1 and total >= 50:
        return "REVIEW_SINGLE_MODEL"
    if max_share >= 0.75 and total >= 50:
        return "REVIEW_MODEL_SKEW"
    if cv >= 1.0 and total >= 30:
        return "REVIEW_HIGH_VARIANCE"
    return "OK"

RISK_MAP = {
    "riba": "RIBA",
    "interest": "RIBA",
    "usury": "RIBA",
    "loan": "RIBA_REVIEW",
    "lending": "RIBA_REVIEW",
    "bond": "RIBA_REVIEW",
    "gambling": "MAYSIR",
    "bet": "MAYSIR",
    "casino": "MAYSIR",
    "speculation": "MAYSIR_REVIEW",
    "derivative": "GHARAR_REVIEW",
    "options": "GHARAR_REVIEW",
    "futures": "GHARAR_REVIEW",
    "leverage": "GHARAR_REVIEW",
    "margin": "GHARAR_REVIEW",
    "swap": "GHARAR_REVIEW",
}

def shariah_flags(term: str) -> str:
    t = term.lower()
    flags = {v for k, v in RISK_MAP.items() if k in t}
    return ",".join(sorted(flags)) if flags else ""

def style_sheet(ws):
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells[: min(len(col_cells), 200)]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

def write_df(ws, df: pd.DataFrame):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    style_sheet(ws)

def add_sheet_text(wb: Workbook, name: str, title: str, lines: list[str]):
    ws = wb.create_sheet(name)
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=16)
    for line in lines:
        ws.append([line])
    ws.column_dimensions["A"].width = 120
    ws.freeze_panes = "A2"
    return ws

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-dir", default=".", help="Input directory")
    ap.add_argument("--out", default="BIZRA_Lexicon_Ledger_v0.2.xlsx", help="Output .xlsx path")
    args = ap.parse_args()

    in_dir = Path(args.in_dir)

    tf = pd.read_csv(in_dir / "bizra_term_frequencies_per_model.csv")
    models = sorted(tf["model"].unique().tolist())
    pivot = tf.pivot_table(index="term", columns="model", values="count", aggfunc="sum", fill_value=0)

    pivot["total"] = pivot.sum(axis=1)
    pivot["mean"] = pivot[models].mean(axis=1)
    pivot["std"] = pivot[models].std(axis=1, ddof=0)
    pivot["cv"] = pivot["std"] / pivot["mean"].replace(0, np.nan)
    pivot["presence_models"] = (pivot[models] > 0).sum(axis=1)
    pivot["entropy"] = pivot.apply(lambda r: entropy(r[models].to_numpy(dtype=float)), axis=1)
    pivot["entropy_norm"] = pivot["entropy"] / math.log(len(models))
    pivot["dominance"] = 1 - pivot["entropy_norm"]
    pivot["max_model_share"] = pivot[models].max(axis=1) / pivot["total"].replace(0, np.nan)

    tf_wide = pivot.reset_index()[["term"] + models + ["total", "mean", "cv", "presence_models", "entropy_norm", "dominance", "max_model_share"]]
    tf_wide = tf_wide.sort_values("total", ascending=False)

    master = tf_wide.copy()
    master["rank_overall"] = master["total"].rank(method="min", ascending=False).astype(int)
    master["drift_action"] = master.apply(lambda r: drift_action(r["total"], int(r["presence_models"]), r["cv"], r["max_model_share"]), axis=1)

    master["category"] = "UNCLASSIFIED"
    master["canonical_term"] = master["term"]
    master["synonyms"] = ""
    master["definition_proposed"] = ""
    master["related_terms"] = ""
    master["status"] = "PROPOSED"
    master["evidence_refs"] = ""
    master["last_updated"] = "2025-12-18"

    for c in ["ihsan_correctness","ihsan_consistency","ihsan_completeness","ihsan_causality","ihsan_ethics","ihsan_evidence"]:
        master[c] = "PENDING"
    master["ihsan_score"] = ""

    master["shariah_risk_flags"] = master["term"].apply(shariah_flags)
    master["shariah_status"] = np.where(master["shariah_risk_flags"] != "", "REVIEW_REQUIRED", "UNKNOWN")
    master["shariah_basis_refs"] = ""
    master["governance_owner"] = "SAT_Ethicist"
    master["change_receipt_id"] = ""
    master["promotion_rule"] = "Evidence_refs>=1 + Ihsan PASS + Shariah REVIEW/PASS"

    cols = ["rank_overall","term","canonical_term","category","synonyms","definition_proposed","related_terms"] + models + [
        "total","mean","cv","presence_models","entropy_norm","dominance","max_model_share","drift_action",
        "status","evidence_refs","last_updated",
        "ihsan_correctness","ihsan_consistency","ihsan_completeness","ihsan_causality","ihsan_ethics","ihsan_evidence","ihsan_score",
        "shariah_status","shariah_risk_flags","shariah_basis_refs","governance_owner","change_receipt_id","promotion_rule"
    ]
    master = master[cols].sort_values("total", ascending=False)

    # delta
    delta_rows = []
    for m in models:
        tmp = master[["term", m, "mean"]].copy()
        tmp["model"] = m
        tmp["delta_vs_mean"] = tmp[m] - tmp["mean"]
        tmp["delta_pct_vs_mean"] = np.where(tmp["mean"] > 0, (tmp["delta_vs_mean"] / tmp["mean"]) * 100, np.nan)
        delta_rows.append(tmp.rename(columns={m: "count"}))
    delta = pd.concat(delta_rows, ignore_index=True)[["term","model","count","mean","delta_vs_mean","delta_pct_vs_mean"]]
    delta = delta.sort_values(["term","delta_pct_vs_mean"], ascending=[True, False])

    alerts = master[(master["drift_action"] != "OK") | (master["shariah_status"] == "REVIEW_REQUIRED")].copy()
    alerts = alerts[["rank_overall","term","total","presence_models","max_model_share","cv","drift_action","shariah_status","shariah_risk_flags","status","evidence_refs","governance_owner"]]
    alerts = alerts.sort_values(["drift_action","total"], ascending=[True, False])

    edges = pd.read_csv(in_dir / "combined_concept_graph_edges.csv")
    with open(in_dir / "combined_concept_graph.json", "r", encoding="utf-8") as f:
        cg = json.load(f)
    nodes = pd.DataFrame(cg["nodes"])
    deg = pd.concat([edges["source"], edges["target"]]).value_counts().rename_axis("id").reset_index(name="degree")
    nodes2 = nodes.merge(deg, on="id", how="left").fillna({"degree": 0})
    nodes2["degree"] = nodes2["degree"].astype(int)
    nodes_view = nodes2[["id","label","weight","degree"]].sort_values("degree", ascending=False)

    themes = pd.read_csv(in_dir / "monthly_theme_trends_all_models.csv")
    months_sorted = sorted(themes["month"].unique())
    last12 = months_sorted[-12:]
    themes12 = themes[themes["month"].isin(last12)].copy()
    themes_pivot = themes12.pivot_table(index="theme", columns="month", values="mentions", aggfunc="mean", fill_value=0).reset_index()

    # metrics summary
    metrics = pd.read_csv(in_dir / "conversation_metrics_all_models.csv")
    summary = metrics.groupby("model").agg(
        conversations=("conversation_id","nunique"),
        total_messages=("messages","sum"),
        total_words=("words","sum"),
        avg_words_per_convo=("words","mean"),
        avg_assistant_to_user_word_ratio=("assistant_to_user_word_ratio","mean"),
        avg_duration_min=("duration_min","mean")
    ).reset_index()

    # governance/dictionary
    gates = pd.DataFrame([
        {"gate":"Evidence Gate (H-stakes)", "rule":"H-stakes terms/policies require >=1 evidence_refs; else BLOCKED_BY_EVIDENCE", "owner":"SAT_Auditor", "proof_artifact":"evidence_refs + receipt hash"},
        {"gate":"Ihsān Gate", "rule":"All 6 checks must be PASS before promotion to CANON", "owner":"SAT_Ethicist", "proof_artifact":"ihsan checks + reviewer signature"},
        {"gate":"Shariah Gate", "rule":"If shariah_risk_flags present → Shariah review required before execution/mint", "owner":"Shariah_Board", "proof_artifact":"review memo + citations"},
        {"gate":"Drift Gate", "rule":"If drift_action != OK → reconciliation required across models", "owner":"Model_Governor", "proof_artifact":"delta report + resolution receipt"},
        {"gate":"Change Control", "rule":"Every term status change must include change_receipt_id", "owner":"SAT_Publisher", "proof_artifact":"signed receipt / tag"},
    ])

    dd = pd.DataFrame([
        {"column":"term", "meaning":"Surface form extracted from chat histories", "notes":"Prefer canonical_term for stable naming"},
        {"column":"canonical_term", "meaning":"Stable name for governance + routing", "notes":"Use when aliases exist"},
        {"column":"dominance", "meaning":"1 - normalized entropy across models (higher = more model-skewed)", "notes":"Used for drift triage"},
        {"column":"drift_action", "meaning":"Auto-triage: model skew / high variance / single-model presence", "notes":"Human review decides"},
        {"column":"evidence_refs", "meaning":"Evidence IDs/links supporting definition and use", "notes":"Required for CANON promotion"},
        {"column":"ihsan_*", "meaning":"Six Ihsān checks (PASS/FAIL/PENDING)", "notes":"All PASS required for CANON"},
        {"column":"shariah_risk_flags", "meaning":"Heuristic flags (Riba/Gharar/Maysir review triggers)", "notes":"Not a ruling; triggers review"},
        {"column":"shariah_status", "meaning":"UNKNOWN / REVIEW_REQUIRED / PASS / FAIL", "notes":"Set PASS/FAIL only after qualified review"},
        {"column":"change_receipt_id", "meaning":"Receipt/hash/tag proving the change", "notes":"Amānah: replayable governance"},
    ])

    # core terms
    top_freq = master[["term","total","dominance","presence_models"]].head(30).rename(columns={"total":"score"})
    top_freq["basis"] = "frequency"
    top_deg = nodes2.sort_values("degree", ascending=False).head(30)[["label","degree"]].rename(columns={"label":"term","degree":"score"})
    top_deg["dominance"] = np.nan
    top_deg["presence_models"] = np.nan
    top_deg["basis"] = "graph_degree"
    core = pd.concat([top_freq[["term","score","dominance","presence_models","basis"]],
                      top_deg[["term","score","dominance","presence_models","basis"]]], ignore_index=True)

    # fingerprints
    fingerprints = {}
    for m in models:
        fingerprints[m] = tf_wide[["term", m]].sort_values(m, ascending=False).head(50).rename(columns={m:"count"})

    # build workbook
    wb = Workbook()
    wb.remove(wb.active)

    readme_lines = [
        "BIZRA Lexicon Ledger — v0.2",
        "",
        "Goal: a single, auditable vocabulary + concept graph across all model chat histories, so BIZRA can scale without semantic drift.",
        "",
        "Design principles (Ihsān):",
        "• Evidence-bounded: terms promote to CANON only with evidence_refs and review.",
        "• Fail-closed: high-stakes terms/policies require evidence + Shariah gate when flagged.",
        "• Amānah: changes must be receipted (change_receipt_id) so 'what we claim' matches 'what we can prove'.",
    ]
    sape_lines = [
        "Purpose: activate untapped synaptic capacity in LLMs via a multi-tier, precision-guided prompt system that (1) probes rarely-fired circuits; (2) unlocks symbolic–neural hybrids; (3) triggers higher-order abstraction; (4) surfaces logic–creative tension.",
        "",
        "7–3–6–9 DNA: 7 modules · 3 passes · 6 checks · 9 probes",
        "Operational rule: 'No assumptions — only verified excellence.'",
    ]

    add_sheet_text(wb, "README", "README", readme_lines)
    add_sheet_text(wb, "SAPE_Protocol", "SAPE Protocol", sape_lines)

    for name, df in [
        ("Lexicon_Master", master),
        ("Drift_Alerts", alerts),
        ("Term_Frequency_By_Model", tf_wide),
        ("Cross_Model_Delta", delta),
        ("Core_Terms", core),
        ("Concept_Nodes", nodes_view),
        ("Concept_Edges", edges),
        ("Theme_Trends_12m", themes_pivot),
        ("Conversation_Metrics", summary),
        ("Governance_Gates", gates),
        ("Data_Dictionary", dd),
    ]:
        ws = wb.create_sheet(name)
        write_df(ws, df)

    for m, df in fingerprints.items():
        ws = wb.create_sheet(f"Fingerprint_{m}")
        write_df(ws, df)

    out_path = Path(args.out)
    wb.save(out_path)
    print(f"Wrote: {out_path.resolve()}")

if __name__ == "__main__":
    main()
